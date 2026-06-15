"""
Массовый калькулятор цен для ОЗОН (Дальний кластер)
Цена находится итеративно: от цены считается комиссия ОЗОН, цена пересчитывается
вверх, пока не достигнется нужная рентабельность.
"""

import json
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path


class OzonMassCalcWidget(tk.Frame):

    COLOR_BG    = "#0F172A"
    COLOR_PANEL = "#1E293B"
    COLOR_ROW1  = "#0F172A"
    COLOR_ROW2  = "#13203A"
    COLOR_TEXT  = "#E2E8F0"
    COLOR_MUTED = "#94A3B8"
    COLOR_GREEN = "#4ADE80"
    COLOR_RED   = "#F87171"
    COLOR_BLUE  = "#3B82F6"

    STATE_FILE    = (Path(sys.executable).parent if getattr(sys, 'frozen', False)
                     else Path(__file__).parent) / "ozon_state.json"

    TARIFF_PER_L  = 1.9    # руб. за литр
    ACQUIRING_PCT = 1.5    # %

    # (название, объем_л, вознаграждение_%, логистика_дальн)
    DEFAULT_PRODUCTS = [
        ("Аптечка АМУ-1",                        23,  44,  630.20),
        ("Ключница К-148",                        13,  43,  410.55),
        ("Ключница К-20",                          3,  43,  236.90),
        ("Ключница К-40",                         13,  43,  410.55),
        ("Ключница К-60",                         14,  43,  425.50),
        ("Почтовый ящик ЯП-4",                   38,  45, 1010.85),
        ("Почтовый ящик ЯП-5",                   45,  45, 1114.35),
        ("Почтовый ящик ЯП-6",                   50,  45, 1354.70),
        ("Почтовый ящик ЯП-8",                   72,  45, 1829.65),
        ("Сейф СМ 20",                            14,  40,  425.50),
        ("Сейф СМ 20-Э",                          14,  40,  425.50),
        ("Сейф СМ 23",                            35,  40, 1010.85),
        ("Сейф СМ 23-Э",                          34,  40,  874.00),
        ("Сейф СМ 25",                            24,  40,  630.20),
        ("Сейф СМ 25-Э",                          24,  40,  630.20),
        ("Сейф СМ 30",                            37,  40, 1010.85),
        ("Сейф СМ 30-Э",                          37,  40, 1010.85),
        ("Сейф СМ 50",                            59,  40, 1354.70),
        ("Сейф СМ 50-Э",                          59,  40, 1354.70),
        ("Стеллаж металлический 200*100*40 4п",   44,  34, 1114.35),
        ("Стеллаж металлический 200*100*50 4п",   55,  34, 1354.70),
        ("Стеллаж металлический 200*100*40 6п",   65,  34, 1624.95),
        ("Шкаф ШБС 01-МИНИ",                      64,  40, 1624.95),
        ("Шкаф ШБС 01-МИНИ-Т",                    62,  40, 1624.95),
        ("Шкаф бухгалтерский ШБС-01-17",         348,  40, 6492.90),
        ("Шкаф бухгалтерский ШБС-02-17",         348,  40, 6492.90),
        ("Стеллаж металлический СТ-СПЭ 4(4)",    44,  40, 1114.35),
        ("Стеллаж металлический СТ-СПЭ 5(4)",    55,  40, 1354.70),
        ("Стеллаж металлический СТ-СПЭ 4(6)",    65,  40, 1624.95),
        ("Шкаф оружейный ШО-5",                    348,  45, 6492.90),
        ("Шкаф оружейный ШО-1",                     65,  45, 1624.95),
        ("Шкаф оружейный ШО-2",                     42,  45, 1114.35),
        ("Шкаф оружейный ШО-3",                     37,  45, 1010.85),
        ("Шкаф оружейный ШО-4",                     88,  45, 2117.15),
        ("Верстак 1200 0 0",                       56,  45, 1354.70),
        ("Верстак 1400 0 0",                       83,  45, 2117.15),
        ("Стеллаж СТ-СПЭ 10.3 4 п",              37,  34, 1010.85),
        ("Стеллаж СТ-СПЭ 10.3 6 п",              52,  34, 1354.70),
        ("Стеллаж СТ-СПЭ 10.6 4 п",              68,  34, 1624.95),
        ("Стеллаж СТ-СПЭ 10.6 6 п",              95,  34, 2117.15),
    ]

    COLUMNS = ["name", "volume", "reward", "logistics", "sc", "pvz", "ret",
               "cost", "reward_rub", "acquiring_rub", "commission", "price", "net_profit"]
    HEADERS = {
        "name":          "Наименование",
        "volume":        "Объем, л",
        "reward":        "Возн. МП, %",
        "logistics":     "Логист. Дальн",
        "sc":            "СЦ, руб",
        "pvz":           "ПВЗ, руб",
        "ret":           "Возврат, руб",
        "cost":          "Себестоимость",
        "reward_rub":    "Возн. МП, руб",
        "acquiring_rub": "Эквайринг, руб",
        "commission":    "Комиссия ОЗОН",
        "price":         "РЕК. ЦЕНА",
        "net_profit":    "Чистая прибыль",
    }
    WIDTHS = {
        "name": 230, "volume": 72, "reward": 95, "logistics": 122,
        "sc": 65, "pvz": 65, "ret": 80,
        "cost": 120, "reward_rub": 120, "acquiring_rub": 115,
        "commission": 130, "price": 140, "net_profit": 130,
    }
    EDITABLE_COLS = {0, 1, 2, 3, 4, 5, 6, 7}  # индексы 0-based

    # ─────────────────────────── init ──────────────────────────

    def __init__(self, parent):
        super().__init__(parent, bg=self.COLOR_BG)
        self._row_counter  = 0
        self._edit_widget  = None
        self._recalc_after = None
        self._build_ui()

    # ─────────────────────────── UI ────────────────────────────

    def _build_ui(self):
        self._build_toolbar()
        self._build_table()
        if not self._load_state():
            for p in self.DEFAULT_PRODUCTS:
                self._insert_row(*p)

    # ── Toolbar ─────────────────────────────────────────────────

    def _build_toolbar(self):
        bar = tk.Frame(self, bg=self.COLOR_BG)
        bar.pack(fill="x", padx=14, pady=(14, 8))

        tk.Label(bar, text="Желаемая рентабельность, %:",
                 bg=self.COLOR_BG, fg=self.COLOR_TEXT,
                 font=("Arial", 10, "bold")).pack(side="left", padx=(0, 8))

        self.roi_var = tk.StringVar(value="30.0")
        roi_sb = tk.Spinbox(
            bar, from_=0, to=9999, increment=1,
            textvariable=self.roi_var, width=9,
            bg=self.COLOR_PANEL, fg=self.COLOR_GREEN,
            relief="flat", font=("Arial", 11, "bold"),
            insertbackground=self.COLOR_TEXT,
            command=self._schedule_recalc_all,
        )
        roi_sb.pack(side="left", padx=(0, 18))
        roi_sb.bind("<KeyRelease>", lambda e: self._schedule_recalc_all())

        # На macOS tk.Button игнорирует bg → используем Label как кнопку
        def themed_btn(text, cmd, bg, hover):
            lbl = tk.Label(bar, text=text, bg=bg, fg=self.COLOR_TEXT,
                           font=("Arial", 9, "bold"), padx=12, pady=5,
                           cursor="hand2")
            lbl.pack(side="left", padx=3)
            lbl.bind("<Button-1>", lambda e: cmd())
            lbl.bind("<Enter>",    lambda e: lbl.config(bg=hover))
            lbl.bind("<Leave>",    lambda e: lbl.config(bg=bg))
            return lbl

        themed_btn("+ Добавить",          self._add_product,            "#1E3A5F", "#2563EB")
        themed_btn("✏ Изменить",           self._edit_selected,           "#14402E", "#047857")
        themed_btn("✕ Удалить",            self._delete_selected,         "#3B0F0F", "#B91C1C")
        themed_btn("⟳ Пересчитать всё",    self._recalculate_all,         "#1E293B", "#334155")
        themed_btn("🚚 Тарифы логистики",  self._open_logistics_settings, "#2D1B69", "#4C1D95")
        themed_btn("📊 Экспорт в Excel",   self._export_excel,            "#1A3A2A", "#166534")

        # ── Вторая строка: управление себестоимостью ─────────────
        bar2 = tk.Frame(self, bg=self.COLOR_BG)
        bar2.pack(fill="x", padx=14, pady=(0, 6))

        tk.Label(bar2, text="Себестоимость:",
                 bg=self.COLOR_BG, fg=self.COLOR_MUTED,
                 font=("Arial", 9, "bold")).pack(side="left", padx=(0, 8))

        def bar2_btn(text, cmd, bg, hover):
            lbl = tk.Label(bar2, text=text, bg=bg, fg=self.COLOR_TEXT,
                           font=("Arial", 9, "bold"), padx=12, pady=4,
                           cursor="hand2")
            lbl.pack(side="left", padx=3)
            lbl.bind("<Button-1>", lambda _: cmd())
            lbl.bind("<Enter>",    lambda _: lbl.config(bg=hover))
            lbl.bind("<Leave>",    lambda _: lbl.config(bg=bg))

        bar2_btn("📋 Скачать шаблон",    self._export_cost_template, "#3B2800", "#92400E")
        bar2_btn("📂 Загрузить из Excel", self._import_cost_template, "#3B2800", "#92400E")

    # ── Таблица ─────────────────────────────────────────────────

    def _build_table(self):
        frame = tk.Frame(self, bg=self.COLOR_BG)
        frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Mass.Treeview",
                        background=self.COLOR_ROW1,
                        foreground=self.COLOR_TEXT,
                        fieldbackground=self.COLOR_ROW1,
                        rowheight=32,
                        font=("Arial", 9))
        style.configure("Mass.Treeview.Heading",
                        background="#1E293B",
                        foreground="#94A3B8",
                        font=("Arial", 9, "bold"),
                        relief="flat")
        style.map("Mass.Treeview",
                  background=[("selected", "#1E3A5F")],
                  foreground=[("selected", "#FFFFFF")])

        self.tree = ttk.Treeview(frame, columns=self.COLUMNS,
                                 style="Mass.Treeview",
                                 selectmode="extended", show="headings")

        for col in self.COLUMNS:
            self.tree.column(col, width=self.WIDTHS[col], anchor="center", minwidth=48)
            self.tree.heading(col, text=self.HEADERS[col])

        self.tree.tag_configure("row1", background=self.COLOR_ROW1)
        self.tree.tag_configure("row2", background=self.COLOR_ROW2)

        vsb = ttk.Scrollbar(frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._on_double_click)

    # ─────────────────────── строки ─────────────────────────────

    def _next_iid(self) -> str:
        iid = f"r{self._row_counter}"
        self._row_counter += 1
        return iid

    def _insert_row(self, name, volume, reward, logistics,
                    sc=20, pvz=25, ret=15, cost="", recalc=True) -> str:
        iid = self._next_iid()
        n   = len(self.tree.get_children())
        tag = "row1" if n % 2 == 0 else "row2"
        self.tree.insert("", "end", iid=iid, tags=(tag,), values=(
            name,
            str(volume),
            str(reward),
            f"{float(logistics):.2f}",
            str(sc), str(pvz), str(ret),
            str(cost) if cost != "" else "",
            "—", "—", "—", "—", "—",
        ))
        if recalc:
            self._recalculate_iid(iid)
        return iid

    def _recolor(self):
        for i, iid in enumerate(self.tree.get_children()):
            self.tree.item(iid, tags=("row1" if i % 2 == 0 else "row2",))

    # ──────────────── редактирование ячеек ──────────────────────

    def _on_double_click(self, event):
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        col_idx = int(self.tree.identify_column(event.x).lstrip("#")) - 1
        iid     = self.tree.identify_row(event.y)
        if iid and col_idx in self.EDITABLE_COLS:
            self._open_cell_editor(iid, col_idx)

    def _open_cell_editor(self, iid: str, col_idx: int):
        bbox = self.tree.bbox(iid, self.COLUMNS[col_idx])
        if not bbox:
            return
        x, y, w, h = bbox
        current = self.tree.item(iid)["values"][col_idx]

        self._close_editor()

        entry = tk.Entry(self.tree,
                         bg="#1E3A5F", fg=self.COLOR_TEXT,
                         insertbackground=self.COLOR_TEXT,
                         relief="solid", bd=1, font=("Arial", 9))
        entry.insert(0, str(current))
        entry.select_range(0, tk.END)
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        self._edit_widget = entry

        def commit(e=None):
            val = entry.get()
            self._close_editor()
            vals = list(self.tree.item(iid)["values"])
            vals[col_idx] = val
            self.tree.item(iid, values=vals)
            self._recalculate_iid(iid)
            self._save_state()

        entry.bind("<Return>",   commit)
        entry.bind("<KP_Enter>", commit)
        entry.bind("<Tab>",      commit)
        entry.bind("<Escape>",   lambda e: self._close_editor())
        entry.bind("<FocusOut>", commit)

    def _close_editor(self):
        if self._edit_widget and self._edit_widget.winfo_exists():
            self._edit_widget.place_forget()
            self._edit_widget.destroy()
        self._edit_widget = None

    # ──────────────── диалоги управления товарами ───────────────

    def _add_product(self):
        dlg = _ProductDialog(self, title="Добавить товар")
        self.wait_window(dlg)
        if dlg.result:
            self._insert_row(**dlg.result)
            self._recolor()
            self._save_state()

    def _edit_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Нет выбора", "Выберите строку для изменения.")
            return
        iid  = sel[0]
        vals = self.tree.item(iid)["values"]
        prefill = dict(name=str(vals[0]), volume=str(vals[1]), reward=str(vals[2]),
                       logistics=str(vals[3]), sc=str(vals[4]), pvz=str(vals[5]),
                       ret=str(vals[6]), cost=str(vals[7]))
        dlg = _ProductDialog(self, title="Изменить товар", prefill=prefill)
        self.wait_window(dlg)
        if dlg.result:
            r = dlg.result
            new_vals = list(vals)
            new_vals[0] = r["name"]
            new_vals[1] = r["volume"]
            new_vals[2] = r["reward"]
            new_vals[3] = f"{float(r['logistics']):.2f}" if r["logistics"] else "0.00"
            new_vals[4] = r["sc"]
            new_vals[5] = r["pvz"]
            new_vals[6] = r["ret"]
            new_vals[7] = r["cost"]
            self.tree.item(iid, values=new_vals)
            self._recalculate_iid(iid)
            self._save_state()

    def _delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        if not messagebox.askyesno("Удалить", f"Удалить {len(sel)} строк(и)?"):
            return
        for iid in sel:
            self.tree.delete(iid)
        self._recolor()
        self._save_state()

    def _open_logistics_settings(self):
        # Открывает отдельное окно для редактирования стоимости логистики по каждому товару
        dlg = _LogisticsDialog(self, self.tree)
        self.wait_window(dlg)

    # ──────────── себестоимость: шаблон и импорт ────────────────

    def _export_cost_template(self):
        """Выгружает xlsx-шаблон: Наименование + Себестоимость (уже заполнена если есть)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Нет библиотеки", "pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            initialfile="себестоимость_шаблон.xlsx",
            title="Сохранить шаблон себестоимости",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Себестоимость"

        fill_hdr  = PatternFill("solid", fgColor="1E3A5F")
        fill_hint = PatternFill("solid", fgColor="2D1B00")
        fill_row1 = PatternFill("solid", fgColor="0F172A")
        fill_row2 = PatternFill("solid", fgColor="13203A")
        font_hdr  = Font(name="Arial", bold=True,  size=10, color="E2E8F0")
        font_hint = Font(name="Arial", italic=True, size=8,  color="94A3B8")
        font_name = Font(name="Arial", size=9,  color="E2E8F0")
        font_cost = Font(name="Arial", bold=True,  size=10, color="4ADE80")
        font_empty= Font(name="Arial", size=10, color="94A3B8")
        align_c = Alignment(horizontal="center", vertical="center")
        align_l = Alignment(horizontal="left",   vertical="center")
        thin    = Side(style="thin", color="334155")
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Подсказка в первой строке
        ws.merge_cells("A1:B1")
        hint = ws.cell(row=1, column=1,
                       value="Заполните колонку «Себестоимость» и загрузите файл обратно через «Загрузить из Excel»")
        hint.font      = font_hint
        hint.fill      = fill_hint
        hint.alignment = align_l
        ws.row_dimensions[1].height = 18

        # Шапка
        for col, (text, w) in enumerate([("Наименование", 48), ("Себестоимость, руб", 22)], start=1):
            cell = ws.cell(row=2, column=col, value=text)
            cell.font      = font_hdr
            cell.fill      = fill_hdr
            cell.alignment = align_c if col == 2 else align_l
            cell.border    = border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[2].height = 22

        # Строки товаров
        for i, iid in enumerate(self.tree.get_children()):
            vals    = self.tree.item(iid)["values"]
            name    = str(vals[0])
            cost_s  = str(vals[7])
            fill    = fill_row1 if i % 2 == 0 else fill_row2
            row_num = i + 3

            nc = ws.cell(row=row_num, column=1, value=name)
            nc.font = font_name; nc.fill = fill; nc.alignment = align_l; nc.border = border

            try:
                cost_v = float(cost_s.replace(",", ".").replace(" ", "").replace("\xa0", ""))
                cc = ws.cell(row=row_num, column=2, value=cost_v)
                cc.font = font_cost
            except (ValueError, TypeError):
                cc = ws.cell(row=row_num, column=2, value=None)
                cc.font = font_empty
            cc.fill = fill; cc.alignment = align_c; cc.border = border
            cc.number_format = '#,##0.00'
            ws.row_dimensions[row_num].height = 20

        ws.freeze_panes = "A3"

        try:
            wb.save(path)
            messagebox.showinfo("Шаблон сохранён",
                                f"Файл: {Path(path).name}\n\n"
                                "Заполните колонку «Себестоимость» и загрузите файл обратно.")
        except PermissionError:
            messagebox.showerror("Ошибка", "Файл открыт в Excel. Закройте его и повторите.")

    def _import_cost_template(self):
        """Загружает xlsx с себестоимостью, сопоставляет по названию, обновляет таблицу."""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Нет библиотеки", "pip install openpyxl")
            return

        path = filedialog.askopenfilename(
            filetypes=[("Excel файл", "*.xlsx *.xls")],
            title="Выбрать файл с себестоимостью",
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
        except Exception as exc:
            messagebox.showerror("Ошибка открытия", str(exc))
            return

        # Читаем все строки файла в словарь {название.lower() → себестоимость}
        cost_map: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            name_cell = str(row[0]).strip()
            cost_cell = row[1] if len(row) > 1 else None
            if not name_cell or cost_cell is None:
                continue
            try:
                cost_map[name_cell.lower()] = float(str(cost_cell).replace(",", ".").replace(" ", ""))
            except (ValueError, TypeError):
                pass

        if not cost_map:
            messagebox.showwarning("Пусто", "В файле не найдено ни одной строки с себестоимостью.")
            return

        matched, skipped = 0, []
        for iid in self.tree.get_children():
            vals = list(self.tree.item(iid)["values"])
            name_key = str(vals[0]).strip().lower()
            if name_key in cost_map:
                vals[7] = f"{cost_map[name_key]:.2f}"
                self.tree.item(iid, values=vals)
                self._recalculate_iid(iid)
                matched += 1
            else:
                skipped.append(str(vals[0]))

        msg = f"Обновлено товаров: {matched}"
        if skipped:
            msg += f"\n\nНе найдено в таблице ({len(skipped)}):\n" + "\n".join(f"  • {n}" for n in skipped[:10])
            if len(skipped) > 10:
                msg += f"\n  ... и ещё {len(skipped) - 10}"
        messagebox.showinfo("Импорт завершён", msg)
        self._save_state()

    # ─────────────────────── экспорт ────────────────────────────

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Нет библиотеки",
                "Установите openpyxl:\n\npip install openpyxl",
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            initialfile="ozon_prices.xlsx",
            title="Сохранить таблицу цен ОЗОН",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Цены ОЗОН"

        # ── Стили ──────────────────────────────────────────────
        fill_hdr  = PatternFill("solid", fgColor="1E3A5F")   # шапка — тёмно-синий
        fill_row1 = PatternFill("solid", fgColor="0F172A")   # нечётные строки
        fill_row2 = PatternFill("solid", fgColor="13203A")   # чётные строки

        font_hdr  = Font(name="Arial", bold=True,  size=10, color="E2E8F0")
        font_name = Font(name="Arial", bold=False, size=9,  color="E2E8F0")
        font_val  = Font(name="Arial", bold=True,  size=10, color="4ADE80")  # зелёный для цен

        align_c = Alignment(horizontal="center", vertical="center")
        align_l = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        thin = Side(style="thin", color="334155")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Шапка ──────────────────────────────────────────────
        headers = ["№", "Наименование", "Комиссия ОЗОН, руб", "Рек. цена, руб", "Чистая прибыль, руб"]
        col_widths = [5, 45, 22, 20, 22]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = font_hdr
            cell.fill      = fill_hdr
            cell.alignment = align_c if col_idx != 2 else align_l
            cell.border    = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 22

        # ── Данные ─────────────────────────────────────────────
        # Индексы в values: 0=name, 10=commission, 11=price, 12=net_profit
        for row_num, iid in enumerate(self.tree.get_children(), start=2):
            vals       = self.tree.item(iid)["values"]
            name       = str(vals[0])
            commission = str(vals[10])
            price      = str(vals[11])
            net_profit = str(vals[12])

            # Парсим числа (убираем пробелы-разделители тысяч)
            def to_num(s):
                try:
                    return float(s.replace(",", ".").replace("\xa0", "").replace(" ", ""))
                except (ValueError, AttributeError):
                    return s

            fill = fill_row1 if (row_num % 2 == 0) else fill_row2

            def styled(r, c, value, fnt, aln):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font      = fnt
                cell.fill      = fill
                cell.alignment = aln
                cell.border    = border
                return cell

            styled(row_num, 1, row_num - 1,          font_name, align_c)
            styled(row_num, 2, name,                  font_name, align_l)
            styled(row_num, 3, to_num(commission),    font_val,  align_c)
            styled(row_num, 4, to_num(price),         font_val,  align_c)
            styled(row_num, 5, to_num(net_profit),    font_val,  align_c)

            # Числовой формат с разделителем тысяч
            for col in (3, 4, 5):
                ws.cell(row=row_num, column=col).number_format = '#,##0.00'

            ws.row_dimensions[row_num].height = 20

        # ── Замораживаем шапку ─────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Добавляем строку итогов ────────────────────────────
        last = self.tree.get_children()
        if last:
            total_row = ws.max_row + 2
            ws.cell(row=total_row, column=2, value="Итого строк:").font = font_hdr
            ws.cell(row=total_row, column=2).fill = fill_hdr
            ws.cell(row=total_row, column=3, value=len(last)).font = font_hdr
            ws.cell(row=total_row, column=3).fill = fill_hdr
            ws.cell(row=total_row, column=3).alignment = align_c

        try:
            wb.save(path)
            messagebox.showinfo("Готово", f"Файл сохранён:\n{Path(path).name}")
        except PermissionError:
            messagebox.showerror("Ошибка", "Файл открыт в Excel. Закройте его и повторите.")

    # ──────────────────── авто-сохранение состояния ─────────────

    def _save_state(self):
        rows = []
        for iid in self.tree.get_children():
            vals = self.tree.item(iid)["values"]
            rows.append({
                "name":      str(vals[0]),
                "volume":    str(vals[1]),
                "reward":    str(vals[2]),
                "logistics": str(vals[3]),
                "sc":        str(vals[4]),
                "pvz":       str(vals[5]),
                "ret":       str(vals[6]),
                "cost":      str(vals[7]),
            })
        state = {"roi": self.roi_var.get(), "rows": rows}
        try:
            with open(self.STATE_FILE, "w", encoding="utf-8") as f:
                json.dump(state, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_state(self) -> bool:
        try:
            with open(self.STATE_FILE, "r", encoding="utf-8") as f:
                state = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            return False
        self.roi_var.set(state.get("roi", "30.0"))
        for row in state.get("rows", []):
            self._insert_row(
                name=row.get("name", ""),
                volume=row.get("volume", "0"),
                reward=row.get("reward", "40"),
                logistics=row.get("logistics", "0"),
                sc=row.get("sc", "20"),
                pvz=row.get("pvz", "25"),
                ret=row.get("ret", "15"),
                cost=row.get("cost", ""),
            )
        return True

    # ─────────────────────── расчёты ────────────────────────────

    def _schedule_recalc_all(self):
        if self._recalc_after is not None:
            self.after_cancel(self._recalc_after)
        self._recalc_after = self.after(400, self._recalculate_all)

    def _recalculate_all(self):
        self._recalc_after = None
        for iid in self.tree.get_children():
            self._recalculate_iid(iid)
        self._save_state()

    def _recalculate_iid(self, iid: str):
        """
        Рекомендованная цена = себестоимость*(рент./100+1) + общая комиссия.
        Так как часть комиссии — проценты от цены, выражаем аналитически:
        price = (cost*(roi/100+1) + fixed_commission) / (1 - pct_rate)
        """
        try:
            roi = float(self.roi_var.get())
        except ValueError:
            roi = 0.0

        try:
            vals = list(self.tree.item(iid)["values"])

            volume     = self._f(vals[1])
            reward_pct = self._f(vals[2])
            logistics  = self._f(vals[3])
            sc         = self._f(vals[4], 20)
            pvz        = self._f(vals[5], 25)
            ret        = self._f(vals[6], 15)
            cost       = self._f(vals[7])

            if cost <= 0:
                vals[8] = "—"
                vals[9] = "—"
                vals[10] = "—"
                vals[11] = "—"
                vals[12] = "—"
                self.tree.item(iid, values=vals)
                return

            shipment         = volume * self.TARIFF_PER_L
            fixed_commission = sc + pvz + ret + shipment + logistics
            pct_rate         = (reward_pct + self.ACQUIRING_PCT) / 100

            price = (cost * (roi / 100 + 1) + fixed_commission) / (1 - pct_rate)

            acquiring  = price * self.ACQUIRING_PCT / 100
            reward_rub = price * reward_pct / 100
            commission = acquiring + sc + pvz + ret + shipment + logistics + reward_rub
            net_profit = price - cost - commission

            vals[8]  = f"{reward_rub:,.2f}"
            vals[9]  = f"{acquiring:,.2f}"
            vals[10] = f"{commission:,.2f}"
            vals[11] = f"{price:,.2f}"
            vals[12] = f"{net_profit:,.2f}"
            self.tree.item(iid, values=vals)

        except Exception:
            pass

    @staticmethod
    def _f(val, default: float = 0.0) -> float:
        try:
            return float(str(val).replace(",", ".").replace("\xa0", "").replace(" ", "") or default)
        except (ValueError, TypeError):
            return default


# ─────────────── диалог добавления / изменения товара ─────────────────

class _ProductDialog(tk.Toplevel):

    FIELDS = [
        ("name",      "Наименование",           "Новый товар"),
        ("volume",    "Объем, л",               "0"),
        ("reward",    "Вознаграждение МП, %",   "40"),
        ("logistics", "Логистика Дальн, руб",   "0"),
        ("sc",        "Отгрузка СЦ, руб",       "20"),
        ("pvz",       "Доставка ПВЗ, руб",      "25"),
        ("ret",       "Обработка возврата, руб","15"),
        ("cost",      "Себестоимость, руб",      ""),
    ]

    def __init__(self, parent, title="Товар", prefill: dict = None):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.configure(bg="#0F172A")
        self.result  = None
        self._entries = {}
        self._build(prefill or {})
        self.transient(parent)
        self.grab_set()
        self.update_idletasks()
        pw = parent.winfo_rootx() + parent.winfo_width()  // 2
        ph = parent.winfo_rooty() + parent.winfo_height() // 2
        self.geometry(f"+{pw - 170}+{ph - 180}")

    def _build(self, prefill):
        pad = dict(padx=16, pady=5)
        for key, label, default in self.FIELDS:
            row = tk.Frame(self, bg="#0F172A")
            row.pack(fill="x", **pad)
            tk.Label(row, text=label, width=26, anchor="w",
                     bg="#0F172A", fg="#94A3B8",
                     font=("Arial", 9)).pack(side="left")
            var = tk.StringVar(value=prefill.get(key, default))
            tk.Entry(row, textvariable=var, width=18,
                     bg="#1E293B", fg="#E2E8F0",
                     insertbackground="#E2E8F0",
                     relief="flat", font=("Arial", 9)).pack(side="left", padx=(8, 0))
            self._entries[key] = var

        btn_row = tk.Frame(self, bg="#0F172A")
        btn_row.pack(fill="x", padx=16, pady=(12, 16))

        def lbtn(parent, text, cmd, bg, hover):
            lbl = tk.Label(parent, text=text, bg=bg, fg="#E2E8F0",
                           font=("Arial", 9, "bold"), padx=14, pady=6,
                           cursor="hand2")
            lbl.bind("<Button-1>", lambda e: cmd())
            lbl.bind("<Enter>",    lambda e: lbl.config(bg=hover))
            lbl.bind("<Leave>",    lambda e: lbl.config(bg=bg))
            return lbl

        lbtn(btn_row, "Сохранить", self._ok,     "#1E3A5F", "#2563EB").pack(side="right", padx=(6, 0))
        lbtn(btn_row, "Отмена",    self.destroy, "#1E293B", "#334155").pack(side="right")

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())

    def _ok(self):
        self.result = {k: v.get().strip() for k, v in self._entries.items()}
        self.destroy()


# ─────────────── диалог настройки тарифов логистики ──────────────────

class _LogisticsDialog(tk.Toplevel):
    """
    Отдельное окно для редактирования стоимости логистики по каждому товару.
    Читает текущие значения из таблицы, после сохранения записывает обратно
    и запускает пересчёт всех строк.
    """

    BG      = "#0F172A"
    BG_ROW1 = "#0F172A"
    BG_ROW2 = "#13203A"
    BG_HDR  = "#1E293B"
    FG      = "#E2E8F0"
    FG_MUTED = "#94A3B8"
    FG_VAL  = "#4ADE80"

    def __init__(self, parent_widget, tree):
        super().__init__(parent_widget)
        self.title("Тарифы логистики — Дальний кластер")
        self.configure(bg=self.BG)
        self.resizable(True, True)
        self._tree    = tree
        self._parent  = parent_widget
        self._entries = {}   # iid → tk.StringVar с текущим значением логистики
        self._build()
        self.transient(parent_widget)
        self.grab_set()
        self.update_idletasks()
        # Центрируем относительно родительского окна
        pw = parent_widget.winfo_rootx() + parent_widget.winfo_width()  // 2
        ph = parent_widget.winfo_rooty() + parent_widget.winfo_height() // 2
        self.geometry(f"560x640+{pw - 280}+{ph - 320}")

    def _build(self):
        # ── Заголовок ──────────────────────────────────────────────
        tk.Label(self, text="Стоимость логистики по каждому товару",
                 bg=self.BG, fg="#3B82F6",
                 font=("Arial", 11, "bold")).pack(pady=(14, 2), padx=16, anchor="w")
        tk.Label(self,
                 text="Дальний кластер (руб.). Изменения применяются после нажатия «Сохранить».",
                 bg=self.BG, fg=self.FG_MUTED,
                 font=("Arial", 8)).pack(pady=(0, 8), padx=16, anchor="w")

        # ── Прокручиваемый список ──────────────────────────────────
        container = tk.Frame(self, bg=self.BG)
        container.pack(fill="both", expand=True, padx=16)

        canvas = tk.Canvas(container, bg=self.BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        inner  = tk.Frame(canvas, bg=self.BG)
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Прокрутка колесом мыши
        canvas.bind("<MouseWheel>",
                    lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        # ── Шапка таблицы ─────────────────────────────────────────
        hdr = tk.Frame(inner, bg=self.BG_HDR)
        hdr.pack(fill="x", pady=(0, 2))
        tk.Label(hdr, text="#",                   bg=self.BG_HDR, fg=self.FG_MUTED,
                 font=("Arial", 8, "bold"), width=3,  anchor="center").pack(side="left", padx=(8, 0), pady=4)
        tk.Label(hdr, text="Наименование товара", bg=self.BG_HDR, fg=self.FG_MUTED,
                 font=("Arial", 8, "bold"), width=36, anchor="w").pack(side="left", padx=8, pady=4)
        tk.Label(hdr, text="Логистика, руб",      bg=self.BG_HDR, fg=self.FG_MUTED,
                 font=("Arial", 8, "bold"), width=14, anchor="center").pack(side="left", padx=8, pady=4)

        # ── Строки товаров ─────────────────────────────────────────
        for i, iid in enumerate(self._tree.get_children()):
            vals = self._tree.item(iid)["values"]
            name      = str(vals[0])
            logistics = str(vals[3])   # индекс 3 = "Логист. Дальн"

            bg  = self.BG_ROW1 if i % 2 == 0 else self.BG_ROW2
            row = tk.Frame(inner, bg=bg)
            row.pack(fill="x", pady=1)

            # Порядковый номер
            tk.Label(row, text=str(i + 1), bg=bg, fg=self.FG_MUTED,
                     font=("Arial", 8), width=3, anchor="center").pack(side="left", padx=(8, 0), pady=3)

            # Название товара
            tk.Label(row, text=name, bg=bg, fg=self.FG,
                     font=("Arial", 9), width=36, anchor="w").pack(side="left", padx=8, pady=3)

            # Поле ввода тарифа
            var = tk.StringVar(value=logistics)
            entry = tk.Entry(row, textvariable=var, width=12,
                             bg="#1E293B", fg=self.FG_VAL,
                             insertbackground=self.FG,
                             relief="flat", font=("Arial", 9, "bold"),
                             justify="right")
            entry.pack(side="left", padx=8, pady=2)

            self._entries[iid] = var

        # ── Кнопки ────────────────────────────────────────────────
        btn_row = tk.Frame(self, bg=self.BG)
        btn_row.pack(fill="x", padx=16, pady=(10, 16))

        def lbtn(text, cmd, bg, hover):
            lbl = tk.Label(btn_row, text=text, bg=bg, fg=self.FG,
                           font=("Arial", 9, "bold"), padx=14, pady=6,
                           cursor="hand2")
            lbl.bind("<Button-1>", lambda e: cmd())
            lbl.bind("<Enter>",    lambda e: lbl.config(bg=hover))
            lbl.bind("<Leave>",    lambda e: lbl.config(bg=bg))
            return lbl

        lbtn("Сохранить и пересчитать", self._save,    "#1E3A5F", "#2563EB").pack(side="right", padx=(6, 0))
        lbtn("Отмена",                  self.destroy,  "#1E293B", "#334155").pack(side="right")

        self.bind("<Escape>", lambda _: self.destroy())

    def _save(self):
        # Записываем новые значения логистики обратно в таблицу
        for iid, var in self._entries.items():
            try:
                new_val = float(var.get().replace(",", ".").replace(" ", ""))
                vals    = list(self._tree.item(iid)["values"])
                vals[3] = f"{new_val:.2f}"   # индекс 3 = logistics
                self._tree.item(iid, values=vals)
            except (ValueError, TypeError):
                pass
        # Пересчитываем все строки с новыми тарифами
        self._parent._recalculate_all()
        self.destroy()


# ───────────────────────── тест ─────────────────────────────────

def test():
    root = tk.Tk()
    root.title("ОЗОН — Калькулятор цен")
    root.geometry("1600x900")
    root.configure(bg="#0F172A")
    OzonMassCalcWidget(root).pack(fill="both", expand=True)
    root.mainloop()


if __name__ == "__main__":
    test()
